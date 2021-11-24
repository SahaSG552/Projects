# Excel
from openpyxl import load_workbook
wb = load_workbook(filename = 'E:\Работа\РАБОЧИЕ ПРОЕКТЫ\GH\GH Facades\Facades.xlsx')
ws = wb['Data(Armadi)']

for row in ws.iter_rows(min_row=1, max_col=13, max_row=1, values_only=True):
    print(row)